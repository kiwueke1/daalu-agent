"""Bulk inventory parsing for the /operations → Bulk import flow.

Two formats accepted:

* **YAML** — a top-level ``devices:`` list (or a bare list) of records
  with the schema below.
* **Excel (.xlsx)** — first sheet, first row = headers. Column names
  are case-insensitive. Empty rows are ignored.

Schema (per row):

* ``name`` (required)
* ``primary_ip`` (required) — CIDR form, e.g. ``10.0.0.5/24``
* ``transport`` (required) — one of ``linux_ssh``, ``redfish``,
  ``junos``, ``iosxr``, ``eos``
* ``site`` (required) — Nautobot location name (case-insensitive)
* ``device_type`` (required) — Nautobot device-type name
* ``role`` (required) — Nautobot device-role name
* ``platform`` (optional) — Nautobot platform name

The parser produces :class:`ParsedRow` records. Validation against the
tenant's actual Nautobot catalog happens in the router, since that's
where the tenant-scoped HTTP client lives.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

import yaml

VALID_TRANSPORTS = {"linux_ssh", "redfish", "junos", "iosxr", "eos"}

REQUIRED_FIELDS = ("name", "primary_ip", "transport", "site", "device_type", "role")
OPTIONAL_FIELDS = ("platform",)
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS


@dataclass
class ParsedRow:
    """One row from the uploaded file, pre-catalog-resolution.

    ``row_index`` is 1-based and matches what the user sees in their
    editor (line for YAML, sheet row for Excel).
    """

    row_index: int
    name: str
    primary_ip: str
    transport: str
    site: str
    device_type: str
    role: str
    platform: str | None = None
    parse_error: str | None = None


def parse_yaml(data: bytes) -> list[ParsedRow]:
    try:
        doc = yaml.safe_load(data)
    except yaml.YAMLError as e:
        raise ValueError(f"could not parse YAML: {e}") from e
    if doc is None:
        return []
    if isinstance(doc, dict) and "devices" in doc:
        records = doc["devices"]
    elif isinstance(doc, list):
        records = doc
    else:
        raise ValueError(
            "YAML must be a list of device records or a mapping with a 'devices' key"
        )
    if not isinstance(records, list):
        raise ValueError("'devices' must be a list")

    out: list[ParsedRow] = []
    for i, rec in enumerate(records, start=1):
        if not isinstance(rec, dict):
            out.append(
                ParsedRow(
                    row_index=i,
                    name="",
                    primary_ip="",
                    transport="",
                    site="",
                    device_type="",
                    role="",
                    parse_error=f"row {i}: expected a mapping, got {type(rec).__name__}",
                )
            )
            continue
        out.append(_record_to_row(i, rec))
    return out


def parse_excel(data: bytes) -> list[ParsedRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "openpyxl is required for Excel uploads — pip install openpyxl"
        ) from e

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — openpyxl raises many shapes
        raise ValueError(f"could not parse Excel: {e}") from e

    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return []
    headers = [str(c).strip().lower() if c is not None else "" for c in header_row]
    if not any(h in headers for h in REQUIRED_FIELDS):
        raise ValueError(
            f"sheet headers must include at least one of: {', '.join(REQUIRED_FIELDS)}"
        )

    out: list[ParsedRow] = []
    # Sheet row 1 is the header → first data row is sheet row 2.
    for sheet_row, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        rec: dict[str, Any] = {}
        for header, cell in zip(headers, raw, strict=False):
            if header in ALL_FIELDS and cell is not None and str(cell).strip() != "":
                rec[header] = str(cell).strip()
        out.append(_record_to_row(sheet_row, rec))
    return out


def _record_to_row(index: int, rec: dict[str, Any]) -> ParsedRow:
    missing = [f for f in REQUIRED_FIELDS if not rec.get(f)]
    parse_error: str | None = None
    if missing:
        parse_error = f"missing required field(s): {', '.join(missing)}"
    transport = str(rec.get("transport", "")).strip()
    if transport and transport not in VALID_TRANSPORTS:
        msg = (
            f"unknown transport '{transport}' — must be one of: "
            f"{', '.join(sorted(VALID_TRANSPORTS))}"
        )
        parse_error = f"{parse_error}; {msg}" if parse_error else msg
    return ParsedRow(
        row_index=index,
        name=str(rec.get("name", "")).strip(),
        primary_ip=str(rec.get("primary_ip", "")).strip(),
        transport=transport,
        site=str(rec.get("site", "")).strip(),
        device_type=str(rec.get("device_type", "")).strip(),
        role=str(rec.get("role", "")).strip(),
        platform=(str(rec["platform"]).strip() if rec.get("platform") else None),
        parse_error=parse_error,
    )


def parse_upload(filename: str, content_type: str | None, data: bytes) -> list[ParsedRow]:
    """Dispatch by filename extension first, MIME second."""
    lower = (filename or "").lower()
    if lower.endswith((".yaml", ".yml")):
        return parse_yaml(data)
    if lower.endswith(".xlsx"):
        return parse_excel(data)
    if content_type:
        if "yaml" in content_type:
            return parse_yaml(data)
        if "sheet" in content_type or "excel" in content_type:
            return parse_excel(data)
    # Last-ditch: try YAML (it'll fail loudly on Excel bytes).
    return parse_yaml(data)
